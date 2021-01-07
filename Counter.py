from openpyxl import load_workbook
from openpyxl import Workbook
import re


def matcher_slice(keywords, cell_data):
    sen = cell_data.lower()
    for key in keywords:
        if re.search(key, sen):
            return True
    return False


def matcher_split(keywords, cell_data):
    clean_sentance = re.sub(r'[^\w]', ' ', cell_data.lower())
    word_list = clean_sentance.split()
    for key in keywords:
        if key in word_list:
            return True
    return False


class Counter():
    def __init__(self, case_list, file_name):
        self.sheet = (load_workbook(str(case_list))).active
        self.file_name = str(file_name)
        self.output = Workbook()
        self.output.active
        self.output.create_sheet('Non_ValueCAN_cases', 1)
        self.output.create_sheet('ValueCAN_cases', 1)
        self.output.create_sheet('ValueCAN_bench_only', 1)
        self.output.create_sheet('Summary', 1)
    
    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def bench_only(self, cell_data):
        with open('bench_only_kw_single.txt') as bench_sig:
            bench_kw_sig_list = [str(kw) for kw in bench_sig.readline().split(', ')]
        with open('bench_only_kw_double.txt') as bench_dou:
            bench_kw_dou_list = [str(kw) for kw in bench_dou.readline().split(', ')]

        if matcher_slice(bench_kw_dou_list, cell_data[1]) or matcher_split(bench_kw_sig_list, cell_data[1]) or matcher_slice(bench_kw_dou_list, cell_data[2]) or matcher_split(bench_kw_sig_list, cell_data[2]):
            return True
        else:
            return False

    def other_list(self, cell_data):
        with open('keyword_single.txt') as keywords_sig:
            kw_single_list = [str(kw)
                              for kw in keywords_sig.readline().split(', ')]

        with open('keyword_double.txt') as keywords_dou:
            kw_double_list = [str(kw)
                              for kw in keywords_dou.readline().split(', ')]

        if matcher_slice(kw_double_list, cell_data[1]) or matcher_slice(kw_double_list, cell_data[2]) or matcher_split(kw_single_list, cell_data[1]) or matcher_split(kw_single_list, cell_data[2]):
            return True
        else:
            return False


    def counter(self):
        sheet = self.sheet
        current = 0
        total_amount = 0
        bench_amount = 0
        for row in sheet.iter_rows(max_col=4, values_only=True):
            current += 1
            detail = self.cell_data(row)
            print('iterate case {}/1499'.format(str(current)))
            if self.bench_only(detail):
                total_amount += 1
                bench_amount += 1
                self.output['ValueCAN_bench_only'].append(detail)
            elif self.other_list(detail):
                total_amount += 1
                self.output['ValueCAN_cases'].append(detail)
            else:
                self.output['Non_ValueCAN_cases'].append(detail)
        
        self.output['Summary'].append(['total cases required ValueCAN', str(total_amount)])
        self.output['Summary'].append(['Percentage', str(round(total_amount*100/1499, 2))+'%'])
        self.output['Summary'].append(['Number of valueCAN only cases that need to perform on the bench', str(bench_amount)])
        self.output['Summary'].append(['Case that can run on both bench and ecomate', str(total_amount - bench_amount)])
        self.output['Summary'].append(['Case does not required valueCAN', str(1499-total_amount)])

        self.output.save(self.file_name)

        print('ValueCAN-related case: {}'.format(str(total_amount)))
        print('Which is {}%'.format(str(round(total_amount * 100 / 1499, 2))))



k = Counter('MY22_1499.xlsx', 'ValueCAN_report.xlsx')
k.counter()
